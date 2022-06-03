# libraries:
import time
import networkx as nx
import matplotlib.pyplot as plt
import tweepy
import openpyxl
from tweepy import NotFound, Unauthorized, HTTPException, BadRequest, Forbidden, TooManyRequests, TweepyException


class TwitterRelationships():
    def __init__(self):
        # TWITTER AUTHINTECHATION:
        auth = tweepy.OAuthHandler("", "")
        auth.set_access_token("",
                              "")


        self.path = "C://Users//User//Desktop//TheBaltimoreProject//Useful Info//Test9.xlsx"

        self.wb_obj = openpyxl.load_workbook(self.path)

        self.sheet_obj = self.wb_obj["Dataset"]
        self.sheet_obj2 = self.wb_obj["Network"]

        self.api = tweepy.API(auth, wait_on_rate_limit=True)

        # self.getUsersIDs()
        #
        self.getRelationships()

        # self.graph = nx.Graph()
        # self.drawGraph()
        #
        # self.getCentralitie()
        #
        # self.showGraph()
    # Converting names into ids and adding them as fields to the dataset:
    def getUsersIDs(self):
        print("Getting user Ids...\n")
        i = 0
        usernames = []
        screenNid = {}
        row = 1
        for organization in self.sheet_obj["H"]:
            organization = organization.value.strip("\n")
            if organization != "Twitter username":
                usernames.append(organization)
                i = i + 1

            if len(usernames) == 99:
                time.sleep(61)
                users = self.api.lookup_users(screen_name=usernames)

                for user in users:
                    screen_name = user.screen_name.lower()
                    screenNid[screen_name] = user.id

                for username in usernames:
                    # print(username)
                    username = username.lower()
                    try:
                        userid = screenNid[username]
                        # print(screenNid)
                    except KeyError:
                        userid = None
                    if self.sheet_obj["BY"][row].value == None:
                        # print(str(row) + ": " + self.sheet_obj["H"][row].value)
                        self.sheet_obj["BY"][row].value = userid
                    row = row + 1
                self.wb_obj.save(self.path)
                usernames.clear()
                i = 0
        if i > 0:
            time.sleep(61)
            try:
                users = self.api.lookup_users(screen_name=usernames)
                for user in users:
                    screen_name = user.screen_name.lower()
                    screenNid[screen_name] = user.id
                for username in usernames:
                    username = username.lower()
                    try:
                        userid = screenNid[username]
                    except KeyError:
                        userid = None
                    if self.sheet_obj["BY"][row].value == None:
                        # print(str(row) + ": " + self.sheet_obj["H"][row].value)
                        self.sheet_obj["BY"][row].value = userid
                    row = row + 1
                self.wb_obj.save(self.path)
                usernames.clear()
                i = 0
            except:
                x = 1
        print(str(len(screenNid))+" user ids were successfully read.\n")


    # GETTING THE RELATIONSHIP FROM THE TWITTER self.api
    def getRelationships(self):
        i = 0
        sleeptime = 1
        errors = 0
        time.sleep(61)
        row = 0
        if len(self.sheet_obj2['A']) == 1:
            networkRows = 1
        else:
            networkRows = len(self.sheet_obj2['A']) + 1

        organizationLookUp = {}
        for row in range(2, len(self.sheet_obj["BY"]) + 1):
            organizationID = self.sheet_obj.cell(column=77, row=row).value
            if organizationID != None:
                organizationLookUp[organizationID] = self.sheet_obj.cell(column=2, row=row).value
        total = len(self.sheet_obj["BY"])
        for row in range(2, len(self.sheet_obj["BY"]) + 1):
            organizationID = self.sheet_obj.cell(column=77, row=row).value
            if organizationID == None or self.sheet_obj.cell(column=82, row=row).value == 1:
                self.sheet_obj.cell(column=82, row=row).value = 1
            else:
                print("Getting the relationship of '"+organizationLookUp[organizationID]+"' "+str(int(row/total*100))+"%\n")
                friends = []
                try:
                    for page in tweepy.Cursor(self.api.get_friend_ids, user_id=organizationID, count=5000).pages():
                        friends.extend(page)
                        time.sleep(60)
                    for friendID in friends:
                        try:
                            # print(organizationLookUp[organizationID] + "$" + organizationLookUp[friendID])
                            self.sheet_obj2.cell(column = 1, row = networkRows).value = organizationLookUp[organizationID]
                            self.sheet_obj2.cell(column = 2, row = networkRows).value = organizationLookUp[friendID]
                            networkRows = networkRows + 1
                        except KeyError:
                            x = 1
                        i = i + 1
                    self.sheet_obj.cell(column=82, row=row).value = 1
                    self.wb_obj.save(self.path)
                except NotFound:
                    self.sheet_obj.cell(column=82, row=row).value = 1
                    self.wb_obj.save(self.path)
                    time.sleep(61)
                    errors = errors + 1
                except Unauthorized:
                    self.sheet_obj.cell(column=82, row=row).value = 1
                    self.wb_obj.save(self.path)
                    time.sleep(61)
                    errors = errors + 1
                except BadRequest:
                    self.sheet_obj.cell(column=82, row=row).value = 1
                    self.wb_obj.save(self.path)
                    time.sleep(61)
                    errors = errors + 1
                except Forbidden:
                    self.sheet_obj.cell(column=82, row=row).value = 1
                    self.wb_obj.save(self.path)
                    time.sleep(61)
                    errors = errors + 1
                except TooManyRequests:
                    self.sheet_obj.cell(column=82, row=row).value = 0
                    self.wb_obj.save(self.path)
                    time.sleep(61)
                    errors = errors + 1
                except HTTPException:
                    self.sheet_obj.cell(column=82, row=row).value = 1
                    self.wb_obj.save(self.path)
                    time.sleep(61)
                    errors = errors + 1
                except TweepyException:
                    self.sheet_obj.cell(column=82, row=row).value = 1
                    self.wb_obj.save(self.path)
                    time.sleep(61)
                    errors = errors + 1
        print("Getting relationships was completed.\n")

    # Showing self.graph and getting centralities
    def drawGraph(self):
        print("Drawing the network graph...\n")
        for row in range(1,len(self.sheet_obj2['A'])+1):
            try:
                node1 = self.sheet_obj2.cell(column=1, row=row).value
                node2 = self.sheet_obj2.cell(column=2, row=row).value
                self.graph.add_edge(node1, node2, weight=0.5)
            except ValueError:
                x = 1
        pos = nx.spring_layout(self.graph, k=0.6, iterations=20)
        nx.draw_networkx(self.graph, pos=pos, with_labels=False)
        print("Drawing the network graph was completed.\n")
    def showGraph(self):
        print("Shwoing the graph...\n")
        plt.show()


    # # #getting the degree centrality of every organization:
    def getCentralitie(self):
        print("Getting the centralities...\n")
        degree = nx.degree_centrality(self.graph)
        betweennes = nx.betweenness_centrality(self.graph)
        closennes = nx.closeness_centrality(self.graph)
        eigen = nx.eigenvector_centrality(self.graph)
        # print(eigen)
        # page = nx.pagerank(self.graph)
        # MAKE SURE ITS IN ORDER OR FIX IT

        for row in range(2, len(self.sheet_obj['B'])+1):
            orgName = self.sheet_obj.cell(column=2, row=row).value
            try:
                self.sheet_obj.cell(column=78, row=row).value = degree[orgName]
                self.sheet_obj.cell(column=79, row=row).value = betweennes[orgName]
                self.sheet_obj.cell(column=80, row=row).value = closennes[orgName]
                self.sheet_obj.cell(column=81, row=row).value = eigen[orgName]
            except KeyError:
                self.sheet_obj.cell(column=78, row=row).value = 0
                self.sheet_obj.cell(column=79, row=row).value = 0
                self.sheet_obj.cell(column=80, row=row).value = 0
                self.sheet_obj.cell(column=81, row=row).value = 0
        self.wb_obj.save(self.path)
        print("Getting the centralities done.\n")


        # for node in self.graph.nodes:
        #     print(node + "$" + str(degree[node]) + "$" + str(betweennes[node]) + "$" + str(closennes[node]) + "$" + str(
        #         eigen[node]))



# zip code analysis. There is a mistake: one organization is missing from each zip code!

# zipCodesDegree = {}
# zipCodesBetween = {}
# zipCodesClose = {}
# zipCodesEign = {}
# # zipCodesPage = {}
# zipCodesCount = {}
# for line in open("zipCodes.txt", "r"):
#     line = line.strip("\n")
#     zipCodesDegree[line] = 0
#     zipCodesBetween[line] = 0
#     zipCodesClose[line] = 0
#     zipCodesEign[line] = 0
# #     zipCodesPage[line] = 0
#     zipCodesCount[line] = 0
# for line in open("degree.txt","r"):
#     line = line.strip("\n")
#     orgdegree = line.split("$")
#     for info in open("OrganizationZipCodes.txt","r"):
#         info = info.strip("\n")
#         orgzip = info.split("$")
#         if orgdegree[0] == orgzip[0]: #add the case where the zip code is not one of the needed zip codes
#             zipCodesCount[orgzip[1]] = zipCodesCount[orgzip[1]] + 1
#             zipCodesDegree[orgzip[1]] = zipCodesDegree[orgzip[1]] + float(orgdegree[1])
#             zipCodesBetween[orgzip[1]] = zipCodesBetween[orgzip[1]] + float(orgdegree[2])
#             zipCodesClose[orgzip[1]] = zipCodesClose[orgzip[1]] + float(orgdegree[3])
#             zipCodesEign[orgzip[1]] = zipCodesEign[orgzip[1]] + float(orgdegree[4])
#             # zipCodesPage[orgzip[1]] = zipCodesPage[orgzip[1]] + float(orgdegree[5])
# print("Zip Codes, Average Degree, Average Betweenness, Average Closeness, Average EigenValues, Number of Organizations")
# for zip in zipCodesDegree:
#     if zipCodesCount[zip] != 0:
#         zipCodesDegree[zip] = zipCodesDegree[zip] / zipCodesCount[zip]
#         print(zip+", "+str(zipCodesDegree[zip])+", "+str(zipCodesBetween[zip])+", "+str(zipCodesClose[zip])+", "+str(zipCodesEign[zip])+", "+str(zipCodesCount[zip]))
#     else:
#         print(zip + ", 0, 0, 0, 0, 0")

